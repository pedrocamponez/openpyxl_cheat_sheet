# from openpyxl import Workbook, load_workbook

# path = r'file path'
# wb = load_workbook(path)
# ws = wb.active

# wb.save('workbook_name.xlsx') -> to save the workbook
# wb.append('[]') -> to create rows in the current sheet
# wb.create_sheet("Sheet_name") -> to create new sheets

# for row in range(1, 11):
#     for col in range(1, 5):
#         char = get_column_letter(col)
#         print(ws[char + str(row)].value)

# ws.merge_cells("A1:D1") -> to merge the cells from A1 to D1 ("range")
# ws.move_range("C1:D11", rows=2, cols=-2) -> to move cells (negative up/left, positive down/right)

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell # -> class to use the tuple[Cell] and cell.value
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook: Workbook = load_workbook(WORKBOOK_PATH)

sheet_name = 'My Sheet'
worksheet: Worksheet = workbook[sheet_name]

row: tuple[Cell]
for row in worksheet.iter_rows(min_row=2): # -> to discard the header
    for cell in row:
        print(cell.value, end='\t') # -> end t is used to show readable data in command line

        if cell.value == 'Maria':
            worksheet.cell(cell.row, 2, 23) # -> changing some data (cell.row, col, new_value)
    print()

# worksheet['B3'].value = 14

workbook.save(WORKBOOK_PATH)